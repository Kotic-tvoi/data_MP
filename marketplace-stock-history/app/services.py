from __future__ import annotations

import csv
import gzip
import json
import shutil
from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.clients import OzonClient, WBClient
from app.config import Settings, get_settings
from app.db import (
    all_report_items, checkpoint, delete_expired_reports, delete_report, fail_report,
    finish_report, get_report, init_db, oldest_deletable_report, start_report, warehouse_summary,
)


def collect_marketplace(marketplace: str, settings: Settings | None = None) -> dict:
    cfg = settings or get_settings()
    init_db(cfg)
    report_id = start_report(marketplace, cfg)
    try:
        client = WBClient(cfg) if marketplace == "wb" else OzonClient(cfg)
        rows, raw_pages = client.fetch()
        stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        raw_path = cfg.raw_dir / marketplace / f"{stamp}_report_{report_id}.json.gz"
        raw_path.parent.mkdir(parents=True, exist_ok=True)
        with gzip.open(raw_path, "wt", encoding="utf-8") as fh:
            json.dump(raw_pages, fh, ensure_ascii=False, separators=(",", ":"))
        finish_report(report_id, rows, str(raw_path), cfg)
        return get_report(report_id, cfg) or {"id": report_id, "status": "success"}
    except Exception as exc:
        fail_report(report_id, exc, cfg)
        raise


def collect_all(settings: Settings | None = None) -> dict[str, dict]:
    cfg = settings or get_settings()
    result: dict[str, dict] = {}
    for marketplace in ("wb", "ozon"):
        try:
            result[marketplace] = collect_marketplace(marketplace, cfg)
        except Exception as exc:
            result[marketplace] = {"status": "failed", "error": str(exc)}
    cleanup(cfg)
    return result


DETAIL_HEADERS = [
    ("marketplace", "Маркетплейс"), ("fulfillment_type", "Схема"),
    ("product_id", "ID товара"), ("sku", "SKU"), ("offer_id", "Артикул продавца"),
    ("vendor_code", "Vendor code"), ("barcode", "Штрихкод"), ("product_name", "Название"),
    ("size_name", "Размер"), ("warehouse_id", "ID склада"), ("warehouse_name", "Склад"),
    ("region_name", "Регион"), ("available", "Доступно, шт."), ("reserved", "Резерв, шт."),
    ("in_way_to_customer", "В пути к клиенту, шт."), ("in_way_from_customer", "Возврат в пути, шт."),
]
SUMMARY_HEADERS = [
    ("warehouse_id", "ID склада"), ("warehouse_name", "Склад"), ("region_name", "Регион"),
    ("fulfillment_type", "Схема"), ("positions_count", "Позиций"), ("available", "Доступно, шт."),
    ("reserved", "Резерв, шт."), ("in_way_to_customer", "В пути к клиенту, шт."),
    ("in_way_from_customer", "Возврат в пути, шт."),
]


def export_report(report_id: int, fmt: str, settings: Settings | None = None) -> Path:
    cfg = settings or get_settings()
    report = get_report(report_id, cfg)
    if not report:
        raise KeyError(f"Отчёт {report_id} не найден")
    rows = all_report_items(report_id, cfg)
    summary = warehouse_summary(report_id, cfg)
    stamp = (report.get("collected_at") or report.get("started_at") or "report").replace(":", "-")
    base = cfg.export_dir / f"{report['marketplace']}_{stamp}_report_{report_id}"
    fmt = fmt.lower()
    if fmt == "json":
        path = base.with_suffix(".json")
        path.write_text(json.dumps({"report": report, "warehouse_summary": summary, "items": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
        return path
    if fmt == "csv":
        path = base.with_suffix(".csv")
        with path.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=[k for k, _ in DETAIL_HEADERS], extrasaction="ignore", delimiter=";")
            writer.writerow({k: title for k, title in DETAIL_HEADERS})
            writer.writerows(rows)
        return path
    if fmt != "xlsx":
        raise ValueError("Поддерживаются xlsx, csv и json")
    path = base.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары по складам"
    ws.append([title for _, title in DETAIL_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(key, "") for key, _ in DETAIL_HEADERS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws2 = wb.create_sheet("Сводка по складам")
    ws2.append([title for _, title in SUMMARY_HEADERS])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in summary:
        ws2.append([row.get(key, "") for key, _ in SUMMARY_HEADERS])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for sheet in (ws, ws2):
        for col in sheet.columns:
            length = min(max(len(str(c.value or "")) for c in col) + 2, 50)
            sheet.column_dimensions[col[0].column_letter].width = length
    wb.save(path)
    return path


def directory_size(path: Path) -> int:
    return sum(p.stat().st_size for p in path.rglob("*") if p.is_file())


def unlink(path: str | Path | None) -> None:
    if not path:
        return
    try:
        Path(path).unlink(missing_ok=True)
    except OSError:
        pass


def cleanup(settings: Settings | None = None) -> dict:
    cfg = settings or get_settings()
    init_db(cfg)
    cutoff = datetime.now(UTC) - timedelta(days=cfg.retention_days)
    removed_raw = delete_expired_reports(cutoff, cfg)
    for path in removed_raw:
        unlink(path)
    removed_files = 0
    for folder in (cfg.raw_dir, cfg.export_dir):
        for file in folder.rglob("*"):
            if file.is_file() and datetime.fromtimestamp(file.stat().st_mtime, UTC) < cutoff:
                unlink(file)
                removed_files += 1
    emergency_deleted = 0
    max_bytes = int(cfg.max_data_size_gb * 1024**3)
    min_free = int(cfg.min_free_space_gb * 1024**3)
    while directory_size(cfg.data_dir) > max_bytes or shutil.disk_usage(cfg.data_dir).free < min_free:
        report = oldest_deletable_report(cfg)
        if not report:
            break
        unlink(delete_report(report["id"], cfg))
        emergency_deleted += 1
    checkpoint(cfg)
    return {
        "retention_days": cfg.retention_days,
        "expired_reports_deleted": len(removed_raw),
        "expired_files_deleted": removed_files,
        "emergency_reports_deleted": emergency_deleted,
        "data_size_bytes": directory_size(cfg.data_dir),
        "free_space_bytes": shutil.disk_usage(cfg.data_dir).free,
    }
